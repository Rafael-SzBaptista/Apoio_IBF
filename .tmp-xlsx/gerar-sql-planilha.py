"""Gera SQL a partir da planilha Apoio 2026 (exceto Oficina Conecte)."""
from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl

SRC = Path(r"c:\Users\User\Downloads\Apoio 2026.xlsx")
OUT = Path(r"c:\Users\User\Documents\church-organizer-main\church-organizer-main\supabase\sql\planilha-apoio-2026.sql")


def esc(value: object) -> str:
    if value is None:
        return "NULL"
    text = str(value).strip().replace("\r\n", "\n").replace("\r", "\n")
    if text == "":
        return "NULL"
    return "'" + text.replace("'", "''") + "'"


def num(value: object) -> str:
    if value is None:
        return "NULL"
    try:
        n = float(value)
    except (TypeError, ValueError):
        return "NULL"
    if n != n:
        return "NULL"
    if n == int(n):
        return str(int(n))
    return repr(n)


def qty_note(value: object) -> tuple[str, str]:
    if value is None or value == "":
        return "NULL", "NULL"
    if isinstance(value, (int, float)):
        return num(value), "NULL"
    text = str(value).strip()
    try:
        return num(float(text.replace(",", "."))), "NULL"
    except ValueError:
        return "NULL", esc(text)


wb = openpyxl.load_workbook(SRC, data_only=True)
# Índices: 0 escala 2º, 1 escala 1º, 2 receitas, 3 precificação, 4 inventário, 5 caixa, 6 oficina (ignorar)
lines: list[str] = []
lines.append("-- Dados da planilha Apoio 2026 (sem a aba Oficina Conecte).")
lines.append("-- Rode no SQL Editor DEPOIS de supabase/sql/rodar-no-editor.sql")
lines.append("-- Substitui estoque, cardápio, preços e programações de exemplo.")
lines.append("-- Não apaga members nem user_roles (o admin já está vinculado).")
lines.append("")
lines.append("DELETE FROM public.event_inventory;")
lines.append("DELETE FROM public.event_assignments;")
lines.append("DELETE FROM public.event_tasks;")
lines.append("DELETE FROM public.finance_entries;")
lines.append("DELETE FROM public.events;")
lines.append("DELETE FROM public.menu_ingredients;")
lines.append("DELETE FROM public.ingredient_prices;")
lines.append("DELETE FROM public.menus;")
lines.append("DELETE FROM public.inventory_items;")
lines.append("")

# --- Inventário ---
inv = wb.worksheets[4]
sector_left = "Papelaria"
sector_right = "Iluminação"
items: list[tuple[str, str, object]] = []
for r in range(5, inv.max_row + 1):
    sl = inv.cell(r, 1).value
    if sl:
        sector_left = str(sl).strip()
    name_l = inv.cell(r, 2).value
    qty_l = inv.cell(r, 5).value
    if name_l:
        items.append((sector_left, str(name_l).strip(), qty_l))
    sr = inv.cell(r, 7).value
    if sr:
        sector_right = str(sr).strip()
    name_r = inv.cell(r, 8).value
    qty_r = inv.cell(r, 11).value
    if name_r:
        items.append((sector_right, str(name_r).strip(), qty_r))

lines.append("-- Inventário")
for sector, name, qty in items:
    q, note = qty_note(qty)
    lines.append(
        "INSERT INTO public.inventory_items (sector, name, quantity, quantity_note) VALUES ("
        f"{esc(sector)}, {esc(name)}, {q}, {note});"
    )
lines.append("")

# --- Preços ---
prec = wb.worksheets[3]
lines.append("-- Tabela de preços (aba precificação, colunas A-E)")
seen_price = set()
for r in range(5, 50):
    name = prec.cell(r, 1).value
    pack = prec.cell(r, 3).value
    unit = prec.cell(r, 4).value
    price = prec.cell(r, 5).value
    if not name or price is None:
        continue
    key = str(name).strip().lower()
    if key in seen_price:
        continue
    seen_price.add(key)
    unit_sql = esc(unit) if unit else esc("un")
    lines.append(
        "INSERT INTO public.ingredient_prices (name, pack_quantity, unit, price, where_to_buy) VALUES ("
        f"{esc(str(name).strip())}, {num(pack)}, {unit_sql}, {num(price)}, NULL)"
        " ON CONFLICT (name) DO UPDATE SET pack_quantity = EXCLUDED.pack_quantity, unit = EXCLUDED.unit, price = EXCLUDED.price;"
    )
lines.append("")

# --- Menus ---
menus = [
    (
        "Bebidas",
        "Copos e bebidas do Conecte.",
        1.48,
        None,
        [
            ("Copos descartáveis", "1,5 und", "Atacadão", None),
            ("Bebida", "200ml", "Atacadão, Dalben", "Distribuição: Coca zero 0,55 · Coca 0,15 · Sprite 0,15 · sucos 0,15"),
        ],
    ),
    (
        "Strogonoff de frango",
        "Strogonoff com arroz, batata palha e bebidas.",
        20,
        25,
        [
            ("Batata palha", "30g", "Atacadão ou Dalben", None),
            ("Arroz", "220g (81,5g cru)", None, None),
            ("Peito de frango", "250g (320g cru)", "Açougue do Concon ou Swift", "Pegar já moído para strogonoff"),
            ("Molho", "70g", "Fazer (Atacadão ou Dalben)", "20g de ketchup + 40g de creme de leite"),
            ("Cebola", "10g", "Atacadão ou Dalben", "Para refogar com o frango"),
            ("Alho", "8g (1 dente médio)", None, "Para temperar o arroz"),
            ("Azeite e sal", "Suficiente", "Armário da tenda", "Temos guardado no armário da tenda"),
            ("Bebidas", "250ml", "Atacadão ou Dalben", None),
            ("Prato descartável", "1 und", None, None),
        ],
    ),
    (
        "Lanche natural",
        "Lanche natural + suco. Servir cortado ao meio.",
        15,
        15,
        [
            ("Lanche natural", "3 metades", "Atacadão ou Dalben", "Servir cortados ao meio"),
            ("Queijo mussarela", "1 fatia / und", None, None),
            ("Alface", "1 folha / und", None, None),
            ("Tomate", "1 rodela / und", None, None),
            ("Pão caseirinho", "1 und", None, None),
            ("Maionese", "8g / und", None, None),
            ("Suco de laranja natural", "310ml", None, None),
        ],
    ),
    (
        "Coffee break",
        "Bebidas, café, bolo e lanche.",
        8,
        15,
        [
            ("Bebidas", "200ml", "Atacadão ou Dalben", "Uva 0,4 · laranja 0,35 · chá gelado 0,1 · água saborizada 0,15"),
            ("Café", "50ml", None, "Levar chaleira elétrica para fazer na hora"),
            ("Bolo", "1 fatia", "Casa de bolos", "Cenoura 0,6 · chocolate 0,4"),
            ("Lanche", "1 und", "Atacadão ou Dalben", "Metades: pão de forma, presunto e queijo"),
        ],
    ),
    (
        "Chocolate quente",
        "Chocolate quente para noite de jogos / pipoca.",
        None,
        None,
        [
            ("Água", "200ml", "Atacadão ou Dalben", None),
            ("Leite em pó", "40g", None, None),
            ("Chocolate em pó 50%", "30g", None, None),
            ("Leite condensado", "20g", None, None),
            ("Amido de milho", "10g", None, None),
        ],
    ),
    (
        "Cachorro quente",
        "Cachorro quente + bebidas. Cada um monta o seu.",
        15,
        20,
        [
            ("Cachorro quente", "2 und", "Atacadão ou Dalben", "Cada um monta o seu"),
            ("Pão para cachorro quente", "1 und / un", None, None),
            ("Salsicha", "1 und / un", None, None),
            ("Purê de batata", "90g (60,2g cru)", None, None),
            ("Ketchup", "15g / un", None, None),
            ("Mostarda", "15g / un", None, None),
            ("Batata palha", "15g / un", None, None),
            ("Bebidas", "250ml", None, None),
            ("Prato descartável", "1 und", None, None),
        ],
    ),
    (
        "Pizza",
        "Pizza + bebidas. Pizza Mais (19) 3289-0320 ou Nonna Braçuda (19) 99585-7744. No Pizza Mais demora até 1h10.",
        25,
        25,
        [
            ("Pizza", "3 pedaços", "Pizza Mais ou Nonna Braçuda", "No Pizza Mais demora até 1h10. Frango c/ catupiry 0,3 · lombinho 0,2 · marguerita 0,2"),
            ("Bebidas", "250ml", None, None),
        ],
    ),
    (
        "Macarronada",
        "Macarronada + bebidas. Cada um monta o seu.",
        15,
        25,
        [
            ("Macarrão", "450g (184g cru)", "Atacadão ou Dalben", "Cada um monta o seu"),
            ("Molho de tomate", "70g", None, None),
            ("Molho branco", "70g", None, None),
            ("Carne moída", "50g", None, None),
            ("Calabresa", "30g", None, None),
            ("Queijo parmesão ralado", "15g", None, None),
            ("Alho", "8g (1 dente médio)", None, None),
            ("Azeite e sal", "-", "Armário da tenda", "Temos guardado no armário da tenda"),
            ("Bebidas", "250ml", None, None),
            ("Prato descartável", "1 und", None, None),
        ],
    ),
    (
        "Hambúrguer",
        "Hambúrguer + bebidas. Carne no Açougue dos Concons.",
        15,
        25,
        [
            ("Hambúrguer", "1 und", "Açougue dos Concons (carne); Atacadão ou Dalben (restante)", None),
            ("Carne", "1 und / un", None, None),
            ("Queijo mussarela", "2 fatias / un", None, None),
            ("Alface", "2 folhas / un", None, None),
            ("Tomate", "3 rodelas / un", None, "Um tomate médio rende cerca de 10 rodelas"),
            ("Cebola", "1/3 rodela / un", None, "Uma cebola média rende cerca de 12 rodelas"),
            ("Pão", "1 und / un", None, None),
            ("Ketchup", "15g / un", None, None),
            ("Maionese", "15g / un", None, None),
            ("Bebidas", "250ml", None, None),
            ("Prato descartável", "1 und", None, None),
        ],
    ),
    (
        "Salgadinhos de festa",
        "Salgados + bebidas. Tel. 19 98196-1051. Não emitem nota fiscal — pagar com o dinheiro físico do Conecte.",
        15,
        20,
        [
            ("Salgadinhos", "20 und", "19 98196-1051", "Não emitem nota fiscal. Distribuição: risole carne 0,3 · risole presunto/queijo 0,3 · coxinha 0,2 · bolinha de queijo 0,2"),
            ("Bebidas", "250ml", None, None),
            ("Prato descartável", "1 und", None, None),
        ],
    ),
    (
        "Esfihas",
        "Esfiha fechada + bebida. Art Massas (19) 3289-1185 — encomendar com antecedência.",
        20,
        20,
        [
            ("Esfiha", "5 und", "Art Massas - (19) 3289-1185", "Encomendar com antecedência. Carne 0,4 · calabresa 0,2 · frango 0,4"),
            ("Bebidas", "250ml", None, None),
        ],
    ),
    (
        "Pipoca + chocolate quente",
        "Pipoca salgada com chocolate quente.",
        10,
        15,
        [
            ("Milho para pipoca", "50g", "Atacadão ou Dalben", None),
            ("Óleo", "25ml", None, None),
            ("Sazon pipoca", "0,5 sachê", None, None),
            ("Chocolate quente", "1 porção", None, "Ver receita de chocolate quente"),
        ],
    ),
    (
        "Picolé",
        "Dois picolés + suco. Sergel — Caio Bueno consegue preço privilegiado. Até 20 un R$3,50; a partir de 20 un R$2,50.",
        10,
        15,
        [
            ("Picolé", "2 und", "Sergel", "Chocolate 0,3 · leite ninho 0,3 · flocos 0,15 · morango 0,15 · limão 0,1"),
            ("Suco", "310ml", None, None),
        ],
    ),
    (
        "Sorvete de massa",
        "Sorvete de massa. Sergel — Caio Bueno consegue preço privilegiado.",
        10,
        15,
        [
            ("Sorvete de massa", "330ml", "Sergel", "Chocolate 0,4 · leite ninho 0,4 · flocos 0,2"),
            ("Casquinha de sorvete", "1 und", "Atacadão", None),
            ("Pote térmico descartável", "1 und", None, None),
            ("Colher plástica descartável", "1 und", None, None),
        ],
    ),
    (
        "Churrasco",
        "Churrasco dos homens — sem ficha de receita na planilha.",
        None,
        None,
        [],
    ),
    (
        "Lanche comunitário",
        "Festa de encerramento — cardápio a confirmar.",
        None,
        None,
        [],
    ),
]

lines.append("-- Cardápio e receitas")
for name, desc, mini, charged, ings in menus:
    lines.append(
        "INSERT INTO public.menus (name, description, min_price_per_person, charged_price_per_person) VALUES ("
        f"{esc(name)}, {esc(desc)}, {num(mini)}, {num(charged)});"
    )
    for i, (iname, qty, where, notes) in enumerate(ings, start=1):
        where_sql = esc(where) if where else "NULL"
        notes_sql = esc(notes) if notes else "NULL"
        lines.append(
            "INSERT INTO public.menu_ingredients (menu_id, name, qty_per_person, kind, where_to_buy, notes, sort_order) "
            "SELECT id, "
            f"{esc(iname)}, {esc(qty)}, 'ingrediente', {where_sql}, {notes_sql}, {i} "
            f"FROM public.menus WHERE name = {esc(name)};"
        )
lines.append("")

# date, title, time, location, food, deco, aliment, notes, phones
events = [
    ("2026-08-01", "Start Impulse + Conecte", "18:00", "EJ / Sala 8 / Quadra / Tenda", "Hambúrguer", "Beatriz", None, None, None),
    ("2026-08-08", "Noite de jogos", "19:00", "Sala 7 e 8 e quadra", "Salgadinhos de festa", "Rafael B", "Beatriz", "Fornecedor não emite NF — pagar com dinheiro físico do Conecte.", "19 98196-1051"),
    ("2026-08-15", "Informal", None, None, None, None, None, None, None),
    ("2026-08-22", "Churrasco dos homens (conferência das mulheres)", "12:00", "Casa do Dinnout?", "Churrasco", None, None, None, None),
    ("2026-08-29", "Culto Conecte", None, None, "Cachorro quente", "Nicholas", "Carla", None, None),
    ("2026-09-05", "Acampa Conecte", None, None, None, None, None, None, None),
    ("2026-09-12", "Informal", None, None, None, None, None, None, None),
    ("2026-09-19", "Culto Conecte", "19:00", "EJ", "Pizza", "Isabela S", "Rafael", "Pizza Mais demora até 1h10.", "(19) 3289-0320 / Nonna Braçuda (19) 99585-7744"),
    ("2026-09-26", "Planejamento", "19:00", "Sala 7", "Esfihas", "Beatriz", "Isabela L", "Na planilha: Esfihas (picolé?). Encomendar esfiha com antecedência.", "(19) 3289-1185"),
    ("2026-10-03", "Culto Conecte", "19:00", "EJ", "Hambúrguer", "Beatriz", "Rafael B", None, None),
    ("2026-10-10", "Informal", None, None, None, None, None, None, None),
    ("2026-10-17", "Culto Conecte", "19:00", "EJ", "Cachorro quente", "Nicholas", "Carla", None, None),
    ("2026-10-24", "Noite de jogos", None, None, "Salgadinhos de festa", "Rafael", "Isabela S", "Fornecedor não emite NF — pagar com dinheiro físico do Conecte.", "19 98196-1051"),
    ("2026-10-31", "Semana da solidariedade", None, None, None, None, None, None, None),
    ("2026-11-07", "Culto Conecte", None, None, "Pizza", "Isabela L", "Rafael B", "Pizza Mais demora até 1h10.", "(19) 3289-0320 / Nonna Braçuda (19) 99585-7744"),
    ("2026-11-14", "DayCamp com Impulse", None, None, "Eles que mandam", "Beatriz", "Rafael B", "Alimentação fica a cargo do Impulse.", None),
    ("2026-11-21", "Culto Conecte", "19:00", "EJ", "Esfihas", "Nicholas", "Carla", "Encomendar na Art Massas com antecedência.", "(19) 3289-1185"),
    ("2026-11-28", "Festa de encerramento", None, None, "Lanche comunitário?", "Isabela S", "Rafael", "Cardápio a confirmar na planilha.", None),
]

food_to_menu = {
    "Hambúrguer": "Hambúrguer",
    "Salgadinhos de festa": "Salgadinhos de festa",
    "Churrasco": "Churrasco",
    "Cachorro quente": "Cachorro quente",
    "Pizza": "Pizza",
    "Esfihas": "Esfihas",
    "Lanche comunitário?": "Lanche comunitário",
}

lines.append("-- Programações (escala 2º semestre 2026)")
lines.append("-- 1º semestre na planilha está só com o gabarito, sem eventos preenchidos.")
today = datetime.date.today().isoformat()
for date, title, time, loc, food, deco, aliment, notes, phones in events:
    menu = food_to_menu.get(food or "")
    status = "realizada" if date < today else "planejada"
    menu_sql = f"(SELECT id FROM public.menus WHERE name = {esc(menu)})" if menu else "NULL"
    lines.append(
        "INSERT INTO public.events (title, event_date, event_time, location, food_label, menu_id, notes, phones, status) VALUES ("
        f"{esc(title)}, {esc(date)}::date, {esc(time) if time else 'NULL'}, {esc(loc) if loc else 'NULL'}, "
        f"{esc(food) if food else 'NULL'}, {menu_sql}, {esc(notes) if notes else 'NULL'}, "
        f"{esc(phones) if phones else 'NULL'}, {esc(status)});"
    )
    for area, person in (("decoração", deco), ("alimentação", aliment)):
        if not person or str(person).strip() in {"-", ""}:
            continue
        lines.append(
            "INSERT INTO public.event_assignments (event_id, member_id, area) "
            "SELECT e.id, m.id, "
            f"{esc(area)} FROM public.events e JOIN public.members m ON m.full_name = {esc(person)} "
            f"WHERE e.title = {esc(title)} AND e.event_date = {esc(date)}::date "
            "ON CONFLICT (event_id, member_id, area) DO NOTHING;"
        )

lines.append("")
lines.append("INSERT INTO public.event_tasks (event_id, title, sort_order)")
lines.append("SELECT e.id, t.title, t.sort_order FROM public.events e")
lines.append("CROSS JOIN (VALUES")
lines.append("  ('Levantar quantidade de pessoas', 1),")
lines.append("  ('Comprar ingredientes', 2),")
lines.append("  ('Preparar o alimento', 3),")
lines.append("  ('Decorar o local', 4),")
lines.append("  ('Servir', 5),")
lines.append("  ('Limpar e guardar', 6),")
lines.append("  ('Guardar nota fiscal e pedir reembolso', 7)")
lines.append(") AS t(title, sort_order);")
lines.append("")
lines.append("-- Caixa Conecte na planilha está zerado (sem gastos/receitas lançados).")

OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
print("wrote", OUT)
print("inventory", len(items), "prices", len(seen_price), "menus", len(menus), "events", len(events))
